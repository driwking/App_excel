import pandas as pd
import openpyxl as opx
import xlsxwriter as xl
import math

config = {"arquivo":"rows0","novo":"modelo"}
novasColunas = ["id","patente","valor"]
colunasDoArquivos = ["cargo","salario","nome"]

def buscarColunas(cols,file):
    
    dataFrame = opx.load_workbook(file+".xlsx")
    dataFrame1 = dataFrame.active
    posicao = [];valores = []
    
    for row in range(0, dataFrame1.max_row):
        for col in dataFrame1.iter_cols(1,dataFrame1.max_column):
            if(col[row].value != None and not col[row].value in cols):
                continue
            posicao.append(col[row].column)
        break
    
    for row in range(1, dataFrame1.max_row):
        for col in dataFrame1.iter_cols(1,dataFrame1.max_column):
            if(col[row].value == None): break
            if(col[row].column in posicao):
                valores.append(col[row].value)                
    
    dataFrame.close()
    return valores

def gravarDados(cols,values,file,afects = 0):
    
    wb = xl.Workbook(file+".xlsx")
    ws = wb.add_worksheet()
    rowPos = 0
    colPos = 0
    afects = int(afects)
    for col in range(0,len(cols)):
        ws.write(rowPos, col,cols[col])
    
    for row in range(1,afects+1):
        for col in range(0, len(cols)):
            ws.write(row, col, values[col])
            
    wb.close()

if(len(colunasDoArquivos) != len(novasColunas)):  raise Exception("Número de colunas precisam ser igual!")

valoresAserSalvos = buscarColunas(colunasDoArquivos,config.get("arquivo"))
rowsAfects = math.ceil(len(valoresAserSalvos)/len(colunasDoArquivos)) # linhas totais afetadas não vazias na primeira coluna
gravarDados(novasColunas,valoresAserSalvos,config.get("novo"),rowsAfects)

# print(valoresAserSalvos) # discomentar para ver os dados capturados